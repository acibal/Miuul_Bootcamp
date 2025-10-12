#####################################################
# AB Testi ile BiddingYöntemlerinin Dönüşümünün Karşılaştırılması
#####################################################

#####################################################
# İş Problemi
#####################################################

# Facebook kısa süre önce mevcut "maximumbidding" adı verilen teklif verme türüne alternatif
# olarak yeni bir teklif türü olan "average bidding"’i tanıttı. Müşterilerimizden biri olan bombabomba.com,
# bu yeni özelliği test etmeye karar verdi veaveragebidding'in maximumbidding'den daha fazla dönüşüm
# getirip getirmediğini anlamak için bir A/B testi yapmak istiyor.A/B testi 1 aydır devam ediyor ve
# bombabomba.com şimdi sizden bu A/B testinin sonuçlarını analiz etmenizi bekliyor.Bombabomba.com için
# nihai başarı ölçütü Purchase'dır. Bu nedenle, istatistiksel testler için Purchasemetriğine odaklanılmalıdır.

import itertools
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
# !pip install statsmodels
import statsmodels.stats.api as sms
from scipy.stats import ttest_1samp, shapiro, levene, ttest_ind, mannwhitneyu, \
    pearsonr, spearmanr, kendalltau, f_oneway, kruskal
from statsmodels.stats.proportion import proportions_ztest

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', 100)
pd.set_option('display.float_format', lambda x: '%.5f' % x)

#####################################################
# Veri Seti Hikayesi
#####################################################

# Bir firmanın web site bilgilerini içeren bu veri setinde kullanıcıların gördükleri ve tıkladıkları
# reklam sayıları gibi bilgilerin yanı sıra buradan gelen kazanç bilgileri yer almaktadır.Kontrol ve Test
# grubu olmak üzere iki ayrı veri seti vardır. Bu veri setleriab_testing.xlsxexcel’ininayrı sayfalarında yer
# almaktadır. Kontrol grubuna Maximum Bidding, test grubuna AverageBiddinguygulanmıştır.

# impression: Reklam görüntüleme sayısı
# Click: Görüntülenen reklama tıklama sayısı
# Purchase: Tıklanan reklamlar sonrası satın alınan ürün sayısı
# Earning: Satın alınan ürünler sonrası elde edilen kazanç

#####################################################
# Proje Görevleri
#####################################################

######################################################
# AB Testing (Bağımsız İki Örneklem T Testi)
######################################################

# 1. Hipotezleri Kur
# 2. Varsayım Kontrolü
#   - 1. Normallik Varsayımı (shapiro)
#   - 2. Varyans Homojenliği (levene)
# 3. Hipotezin Uygulanması
#   - 1. Varsayımlar sağlanıyorsa bağımsız iki örneklem t testi
#   - 2. Varsayımlar sağlanmıyorsa mannwhitneyu testi
# 4. p-value değerine göre sonuçları yorumla
# Not:
# - Normallik sağlanmıyorsa direkt 2 numara. Varyans homojenliği sağlanmıyorsa 1 numaraya arguman girilir.
# - Normallik incelemesi öncesi aykırı değer incelemesi ve düzeltmesi yapmak faydalı olabilir.

#####################################################
# Görev 1:  Veriyi Hazırlama ve Analiz Etme
#####################################################

# Adım 1:  ab_testing_data.xlsx adlı kontrol ve test grubu verilerinden oluşan veri setini okutunuz. Kontrol ve test grubu verilerini ayrı değişkenlere atayınız.

df_control = pd.read_excel("Data_Science_BC/W4_Measurement/Case_Studies/ABTesti/ab_testing.xlsx", sheet_name="Control Group")
df_test = pd.read_excel("Data_Science_BC/W4_Measurement/Case_Studies/ABTesti/ab_testing.xlsx", sheet_name="Test Group")
df_control.head()
df_test.head()

# Adım 2: Kontrol ve test grubu verilerini analiz ediniz.

df_control.shape
df_test.shape
df_control.info()
df_test.info()
df_control.describe().T
df_test.describe().T

# Adım 3: Analiz işleminden sonra concat metodunu kullanarak kontrol ve test grubu verilerini birleştiriniz.

df_control["before"] = 1
df_control.head()
df_test["before"] = 0
df_test.head()

df = pd.concat([df_control, df_test], axis = 0, ignore_index=True)
df.head(80)
df.shape
df

#####################################################
# Görev 2:  A/B Testinin Hipotezinin Tanımlanması
#####################################################

# Adım 1: Hipotezi tanımlayınız.

# H0 = H1

# Adım 2: Kontrol ve test grubu için purchase(kazanç) ortalamalarını analiz ediniz

df_control.head()
df_test.head()
df_control["Purchase"].mean()
df_control["Purchase"].sum()
df_test["Purchase"].mean()
df_test["Purchase"].sum()
df[df["before"] == 1]["Purchase"].mean()
df[df["before"] == 1]["Purchase"].sum()
df[df["before"] == 0]["Purchase"].mean()
df[df["before"] == 0]["Purchase"].sum()

#####################################################
# GÖREV 3: Hipotez Testinin Gerçekleştirilmesi
#####################################################

######################################################
# AB Testing (Bağımsız İki Örneklem T Testi)
######################################################

# Adım 1: Hipotez testi yapılmadan önce varsayım kontrollerini yapınız.Bunlar Normallik Varsayımı ve Varyans Homojenliğidir.

# Kontrol ve test grubunun normallik varsayımına uyup uymadığını Purchase değişkeni üzerinden ayrı ayrı test ediniz

test_stat, pvalue = shapiro(df.loc[df["before"] == 1, "Purchase"])
print('Test Stat = %.4f, p-value = %.4f' % (test_stat, pvalue))
test_stat, pvalue = shapiro(df.loc[df["before"] == 0, "Purchase"])
print('Test Stat = %.4f, p-value = %.4f' % (test_stat, pvalue))

test_stat, pvalue = levene(df.loc[df["before"] == 1, "Purchase"],
                           df.loc[df["before"] == 0, "Purchase"])
print('Test Stat = %.4f, p-value = %.4f' % (test_stat, pvalue))

# Adım 2: Normallik Varsayımı ve Varyans Homojenliği sonuçlarına göre uygun testi seçiniz

# Hipotez (H0: M1 = M2)
test_stat, pvalue = ttest_ind(df.loc[df["before"] == 1, "Purchase"],
                                 df.loc[df["before"] == 0, "Purchase"])
print('Test Stat = %.4f, p-value = %.4f' % (test_stat, pvalue))

# Adım 3: Test sonucunda elde edilen p_value değerini göz önünde bulundurarak kontrol ve test grubu satın alma
# ortalamaları arasında istatistiki olarak anlamlı bir fark olup olmadığını yorumlayınız.

# test sonucuna göre fark yok. Varsa şans eseri

##############################################################
# GÖREV 4 : Sonuçların Analizi
##############################################################

# Adım 1: Hangi testi kullandınız, sebeplerini belirtiniz.

# mannwhitneyu varsayımlardan en az biri sağlanmadı.

# Adım 2: Elde ettiğiniz test sonuçlarına göre müşteriye tavsiyede bulununuz.



######################################## sonra bak buna #########################

from openpyxl import load_workbook
wb = load_workbook("Measurement Problems/ABTesti/ab_testing.xlsx")
print(wb.sheetnames)

##################################################################################